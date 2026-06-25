"""轻量文本类加载器。"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator

from bs4 import BeautifulSoup
from langchain_core.document_loaders import BaseLoader
from langchain_core.documents import Document


class TextFileLoader(BaseLoader):
    def __init__(self, file_path: str, metadata: dict) -> None:
        self.file_path = file_path
        self.metadata = metadata

    def lazy_load(self) -> Iterator[Document]:
        text = Path(self.file_path).read_text(encoding="utf-8", errors="replace")
        yield Document(page_content=text, metadata=self.metadata.copy())


class HtmlFileLoader(BaseLoader):
    def __init__(self, file_path: str, metadata: dict) -> None:
        self.file_path = file_path
        self.metadata = metadata

    def lazy_load(self) -> Iterator[Document]:
        html = Path(self.file_path).read_text(encoding="utf-8", errors="replace")
        text = BeautifulSoup(html, "html.parser").get_text("\n", strip=True)
        yield Document(page_content=text, metadata=self.metadata.copy())


class CsvFileLoader(BaseLoader):
    def __init__(self, file_path: str, metadata: dict) -> None:
        self.file_path = file_path
        self.metadata = metadata

    def lazy_load(self) -> Iterator[Document]:
        rows: list[str] = []
        with Path(self.file_path).open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append("\t".join(row))
        yield Document(page_content="\n".join(rows), metadata=self.metadata.copy())


class XlsxFileLoader(BaseLoader):
    def __init__(self, file_path: str, metadata: dict) -> None:
        self.file_path = file_path
        self.metadata = metadata

    def lazy_load(self) -> Iterator[Document]:
        from openpyxl import load_workbook

        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                lines: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        lines.append("\t".join(values))
                if lines:
                    metadata = {**self.metadata, "sheet": sheet.title}
                    yield Document(page_content="\n".join(lines), metadata=metadata)
        finally:
            workbook.close()
